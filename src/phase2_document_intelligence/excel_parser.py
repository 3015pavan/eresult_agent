"""
Excel / CSV Parser — Phase 2.

Handles:
  .xlsx / .xls   — openpyxl (primary), xlrd fallback
  .csv           — pandas (primary), csv module fallback

Output: ParsedDocument with text (header + data summary) + tables (raw rows).
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Optional

from .router import ParsedDocument

logger = logging.getLogger(__name__)


def _read_xlsx(path: str) -> list[list[list[str]]]:
    """Read all sheets from an Excel file. Returns list of tables (one per sheet)."""
    try:
        import openpyxl  # type: ignore
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        tables: list[list[list[str]]] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                # Skip completely empty rows
                if all(c is None for c in row):
                    continue
                rows.append([str(c).strip() if c is not None else "" for c in row])
            if rows:
                tables.append(rows)
        return tables
    except ImportError:
        logger.warning("openpyxl not installed, trying xlrd")
    except Exception as exc:
        logger.warning("openpyxl failed: %s", exc)

    # xlrd fallback (old .xls)
    try:
        import xlrd  # type: ignore
        wb = xlrd.open_workbook(path)
        tables = []
        for sheet in wb.sheets():
            rows = []
            for ri in range(sheet.nrows):
                row = [str(sheet.cell_value(ri, ci)).strip() for ci in range(sheet.ncols)]
                if any(row):
                    rows.append(row)
            if rows:
                tables.append(rows)
        return tables
    except ImportError:
        pass
    except Exception as exc:
        logger.warning("xlrd failed: %s", exc)

    return []


def _read_csv(path: str) -> list[list[list[str]]]:
    """Read a CSV file. Returns a single-sheet table."""
    # Try pandas first for better dialect detection
    try:
        import pandas as pd  # type: ignore
        df = pd.read_csv(path, dtype=str, keep_default_na=False)
        rows = [list(df.columns)] + df.values.tolist()
        return [[[str(c).strip() for c in row] for row in rows]]
    except ImportError:
        pass
    except Exception as exc:
        logger.debug("pandas csv read failed: %s", exc)

    # stdlib csv fallback
    try:
        with open(path, encoding="utf-8-sig", errors="replace") as f:
            reader = csv.reader(f)
            rows = [[c.strip() for c in row] for row in reader if any(row)]
        return [rows] if rows else []
    except Exception as exc:
        logger.warning("csv stdlib failed: %s", exc)

    return []


def _tables_to_text(tables: list[list[list[str]]]) -> str:
    """Convert extracted tables to readable text."""
    parts: list[str] = []
    for i, tbl in enumerate(tables):
        if not tbl:
            continue
        header = tbl[0] if tbl else []
        parts.append(f"Sheet/Table {i + 1}: {', '.join(header)}")
        for row in tbl[1:6]:  # First 5 data rows in summary
            parts.append("  " + " | ".join(row))
        if len(tbl) > 7:
            parts.append(f"  ... ({len(tbl) - 1} total rows)")
    return "\n".join(parts)


def parse_spreadsheet(path: str, mime_type: str = "") -> ParsedDocument:
    """
    Parse Excel or CSV file.
    Returns table data in ParsedDocument.tables.
    """
    ext = path.rsplit(".", 1)[-1].lower() if "." in path else ""

    if ext in ("xlsx", "xls") or "spreadsheet" in mime_type or "excel" in mime_type:
        tables = _read_xlsx(path)
        strategy = "openpyxl"
    else:
        tables = _read_csv(path)
        strategy = "pandas_csv"

    if not tables:
        return ParsedDocument(
            source_path=path,
            mime_type=mime_type,
            parse_strategy=f"{strategy}_empty",
            confidence=0.0,
            errors=["no_data_extracted"],
        )

    total_rows = sum(len(t) for t in tables)
    text = _tables_to_text(tables)
    confidence = min(0.95, 0.70 + total_rows * 0.002)

    return ParsedDocument(
        source_path=path,
        mime_type=mime_type or f"application/{ext}",
        text=text,
        tables=tables,
        parse_strategy=strategy,
        confidence=confidence,
        metadata={"sheets": len(tables), "total_rows": total_rows},
    )


def parse_bytes(data: bytes, filename: str) -> ParsedDocument:
    """Parse Excel/CSV from bytes (e.g. attachment content)."""
    import tempfile
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "csv"
    with tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False) as tmp:
        tmp.write(data)
        tmp_path = tmp.name

    try:
        return parse_spreadsheet(tmp_path)
    finally:
        import os
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
